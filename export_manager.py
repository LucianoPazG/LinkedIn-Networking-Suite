#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exportador a Excel para LinkedIn Networking Suite
Exporta contactos, interacciones y recordatorios a Excel
"""

import os
from datetime import datetime
from typing import List, Dict, Optional
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import ContactDatabase

logger = logging.getLogger(__name__)


class ExportManager:
    """Maneja la exportación de datos a Excel"""

    def __init__(self, db: ContactDatabase):
        """
        Inicializa el exportador

        Args:
            db: Instancia de ContactDatabase
        """
        self.db = db

    def export_all_contacts(self, status_filter: Optional[str] = None) -> Optional[str]:
        """
        Exporta todos los contactos a Excel

        Args:
            status_filter: Filtrar por estado (opcional)

        Returns:
            Ruta del archivo o None
        """
        try:
            # Obtener contactos
            contacts = self.db.get_all_contacts(status=status_filter)

            if not contacts:
                print("❌ No hay contactos para exportar")
                return None

            # Convertir a DataFrame
            df = pd.DataFrame(contacts)

            # Renombrar columnas para mejor legibilidad
            column_map = {
                'id': 'ID',
                'linkedin_url': 'LinkedIn URL',
                'name': 'Nombre',
                'job_title': 'Cargo',
                'company': 'Empresa',
                'location': 'Ubicación',
                'industry': 'Industria',
                'about': 'Sobre mí',
                'skills': 'Habilidades',
                'notes': 'Notas',
                'first_contact_date': 'Fecha Primer Contacto',
                'last_contact_date': 'Fecha Último Contacto',
                'status': 'Estado',
                'connection_message_sent': 'Mensaje Enviado',
                'follow_up_count': 'Follow-ups',
                'created_at': 'Creado',
                'updated_at': 'Actualizado'
            }

            df = df.rename(columns=column_map)

            # Seleccionar y ordenar columnas principales
            main_columns = [
                'ID', 'Nombre', 'Cargo', 'Empresa', 'Ubicación',
                'Industria', 'Estado', 'Follow-ups', 'Habilidades',
                'Notas', 'LinkedIn URL'
            ]

            # Filtrar solo columnas que existen
            columns_to_export = [col for col in main_columns if col in df.columns]

            # Agregar columnas de fecha si existen
            for col in ['Fecha Primer Contacto', 'Fecha Último Contacto']:
                if col in df.columns:
                    columns_to_export.append(col)

            df_export = df[columns_to_export]

            # Generar nombre de archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            status_suffix = f"_{status_filter}" if status_filter else ""
            filename = f"contactos_networking{status_suffix}_{timestamp}.xlsx"

            # Crear Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Hoja principal de contactos
                df_export.to_excel(writer, sheet_name='Contactos', index=False)

                # Hoja de estadísticas
                stats = self.db.get_statistics()
                if stats:
                    self._add_statistics_sheet(writer, stats, df_export)

                # Hoja de distribución por estado
                if 'Estado' in df_export.columns:
                    self._add_status_distribution_sheet(writer, df_export)

                # Hoja de top empresas
                if 'Empresa' in df_export.columns:
                    self._add_top_companies_sheet(writer, df_export)

            # Aplicar formato
            self._format_excel(filename)

            print(f"✅ Contactos exportados: {len(contacts)}")
            print(f"📁 Archivo: {filename}")

            return filename

        except Exception as e:
            logger.error(f"Error exportando contactos: {e}")
            import traceback
            traceback.print_exc()
            return None

    def export_contact_interactions(self, contact_id: int) -> Optional[str]:
        """
        Exporta todas las interacciones de un contacto

        Args:
            contact_id: ID del contacto

        Returns:
            Ruta del archivo o None
        """
        try:
            contact = self.db.get_contact(contact_id)

            if not contact:
                print(f"❌ Contacto {contact_id} no encontrado")
                return None

            interactions = self.db.get_contact_interactions(contact_id)

            if not interactions:
                print(f"❌ El contacto no tiene interacciones")
                return None

            # Convertir a DataFrame
            df = pd.DataFrame(interactions)

            # Renombrar columnas
            column_map = {
                'id': 'ID',
                'contact_id': 'ID Contacto',
                'interaction_type': 'Tipo',
                'message': 'Mensaje',
                'outcome': 'Resultado',
                'next_follow_up_date': 'Próximo Follow-up',
                'created_at': 'Fecha'
            }

            df = df.rename(columns=column_map)

            # Seleccionar columnas
            columns_to_export = ['ID', 'Tipo', 'Mensaje', 'Resultado', 'Próximo Follow-up', 'Fecha']
            df_export = df[[col for col in columns_to_export if col in df.columns]]

            # Generar nombre
            name_clean = contact['name'].replace(' ', '_')[:30]
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"interacciones_{name_clean}_{timestamp}.xlsx"

            # Exportar
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_export.to_excel(writer, sheet_name='Interacciones', index=False)

                # Hoja de resumen
                self._add_interaction_summary_sheet(writer, contact, df_export)

            self._format_excel(filename)

            print(f"✅ Interacciones exportadas: {len(interactions)}")
            print(f"📁 Archivo: {filename}")

            return filename

        except Exception as e:
            logger.error(f"Error exportando interacciones: {e}")
            return None

    def export_reminders(self, days_ahead: int = 7) -> Optional[str]:
        """
        Exporta recordatorios pendientes a Excel

        Args:
            days_ahead: Días adelante a incluir

        Returns:
            Ruta del archivo o None
        """
        try:
            reminders = self.db.get_pending_reminders(days_ahead)

            if not reminders:
                print("❌ No hay recordatorios pendientes")
                return None

            # Convertir a DataFrame
            df = pd.DataFrame(reminders)

            # Renombrar columnas
            column_map = {
                'reminder_id': 'ID Recordatorio',
                'reminder_date': 'Fecha',
                'reminder_type': 'Tipo',
                'message': 'Mensaje Sugerido',
                'contact_id': 'ID Contacto',
                'name': 'Nombre',
                'company': 'Empresa',
                'job_title': 'Cargo',
                'linkedin_url': 'LinkedIn URL'
            }

            df = df.rename(columns=column_map)

            # Seleccionar columnas
            columns_to_export = [
                'ID Recordatorio', 'Fecha', 'Nombre', 'Empresa',
                'Cargo', 'Tipo', 'Mensaje Sugerido', 'LinkedIn URL'
            ]
            df_export = df[[col for col in columns_to_export if col in df.columns]]

            # Ordenar por fecha
            df_export = df_export.sort_values('Fecha')

            # Generar nombre
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"recordatorios_{timestamp}.xlsx"

            # Exportar
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_export.to_excel(writer, sheet_name='Recordatorios', index=False)

                # Hoja de resumen
                self._add_reminders_summary_sheet(writer, df_export)

            self._format_excel(filename)

            print(f"✅ Recordatorios exportados: {len(reminders)}")
            print(f"📁 Archivo: {filename}")

            return filename

        except Exception as e:
            logger.error(f"Error exportando recordatorios: {e}")
            return None

    def export_full_report(self) -> Optional[str]:
        """
        Exporta un reporte completo con todo

        Returns:
            Ruta del archivo o None
        """
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"reporte_completo_networking_{timestamp}.xlsx"

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. Contactos
                contacts = self.db.get_all_contacts()
                if contacts:
                    df_contacts = pd.DataFrame(contacts)
                    df_contacts.to_excel(writer, sheet_name='Contactos', index=False)

                # 2. Estadísticas
                stats = self.db.get_statistics()
                if stats:
                    self._add_statistics_sheet(writer, stats, pd.DataFrame(contacts) if contacts else None)

                # 3. Recordatorios
                reminders = self.db.get_pending_reminders(days_ahead=30)
                if reminders:
                    df_reminders = pd.DataFrame(reminders)
                    df_reminders.to_excel(writer, sheet_name='Recordatorios', index=False)

                # 4. Distribución
                if contacts:
                    df_contacts = pd.DataFrame(contacts)
                    if 'status' in df_contacts.columns:
                        self._add_status_distribution_sheet(writer, df_contacts)

                # 5. Top empresas
                if contacts:
                    if 'company' in df_contacts.columns:
                        self._add_top_companies_sheet(writer, df_contacts)

            self._format_excel(filename)

            print(f"✅ Reporte completo exportado")
            print(f"📁 Archivo: {filename}")

            return filename

        except Exception as e:
            logger.error(f"Error exportando reporte completo: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _add_statistics_sheet(self, writer, stats: Dict, df_contacts: Optional[pd.DataFrame] = None):
        """Agrega hoja de estadísticas"""
        stats_data = []

        for key, value in stats.items():
            if isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    stats_data.append({'Métrica': f"{key}_{sub_key}", 'Valor': sub_value})
            else:
                stats_data.append({'Métrica': key, 'Valor': value})

        df_stats = pd.DataFrame(stats_data)
        df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)

    def _add_status_distribution_sheet(self, writer, df: pd.DataFrame):
        """Agrega hoja de distribución por estado"""
        if 'Estado' in df.columns:
            status_dist = df['Estado'].value_counts().reset_index()
            status_dist.columns = ['Estado', 'Cantidad']
            status_dist.to_excel(writer, sheet_name='Distribución Estado', index=False)

    def _add_top_companies_sheet(self, writer, df: pd.DataFrame):
        """Agrega hoja de top empresas"""
        if 'Empresa' in df.columns:
            # Filtrar vacíos
            df_filtered = df[df['Empresa'].notna() & (df['Empresa'] != '')]
            top_companies = df_filtered['Empresa'].value_counts().head(20).reset_index()
            top_companies.columns = ['Empresa', 'Cantidad']
            top_companies.to_excel(writer, sheet_name='Top Empresas', index=False)

    def _add_interaction_summary_sheet(self, writer, contact: Dict, df: pd.DataFrame):
        """Agrega hoja de resumen de interacciones"""
        summary = {
            'Contacto': [contact['name']],
            'Empresa': [contact.get('company', 'N/A')],
            'Cargo': [contact.get('job_title', 'N/A')],
            'Estado': [contact.get('status', 'N/A')],
            'Total Interacciones': [len(df)]
        }

        df_summary = pd.DataFrame(summary)
        df_summary.to_excel(writer, sheet_name='Resumen', index=False)

    def _add_reminders_summary_sheet(self, writer, df: pd.DataFrame):
        """Agrega hoja de resumen de recordatorios"""
        summary = {
            'Total Recordatorios': [len(df)],
            'Para Hoy': [(df['Fecha'].dt.date == datetime.now().date()).sum()],
            'Esta Semana': [len(df)],
            'Por Tipo': df['Tipo'].value_counts().to_dict()
        }

        summary_data = []
        for key, value in summary.items():
            if isinstance(value, dict):
                for type_name, count in value.items():
                    summary_data.append({'Categoría': type_name, 'Valor': count})
            else:
                summary_data.append({'Categoría': key, 'Valor': value})

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Resumen', index=False)

    def _format_excel(self, filename: str):
        """Aplica formato profesional al Excel"""
        try:
            wb = load_workbook(filename)

            # Estilos
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for sheet in wb.worksheets:
                # Ajustar ancho de columnas
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Formatear encabezados
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                # Congelar primera fila
                sheet.freeze_panes = 'A2'

            wb.save(filename)

        except Exception as e:
            logger.error(f"Error formateando Excel: {e}")
